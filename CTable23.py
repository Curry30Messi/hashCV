import csv
import openpyxl
import xlrd
import os
import time
from preprocess_seq import seq_standardize
import random
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import get_column_letter
from collections import Counter


def sort_bases(sequence):
    bases = {'A': False, 'T': False, 'G': False, 'C': False}
    result = []

    for base in sequence:
        if base in bases and not bases[base]:
            result.append(base)
            bases[base] = True

    order = {'A': 0, 'T': 1, 'G': 2, 'C': 3}
    result_sorted = sorted(result, key=lambda x: order[x])
    sorted_result_str = ''.join(result_sorted)

    base_counts = Counter(sequence)
    count_str_list = [f"{base}{base_counts[base]}" if base_counts[base] > 0 else "" for base in 'ATGC']
    count_str = ''.join(count_str_list)

    return sorted_result_str, count_str


def process_table3():
    workbook1 = xlrd.open_workbook("NewSNP.xls")
    sheet1 = workbook1.sheet_by_index(0)

    # 创建新的Excel文件，用于保存复制后的数据
    workbook2 = Workbook()
    sheet2 = workbook2.active
    sheet2.title = 'Sheet2'
    # 从原始表格中复制前两列数据到新表格中
    for row_idx in range(sheet1.nrows):
        row_data = sheet1.row_values(row_idx, end_colx=2)  # 读取指定列范围的数据
        sheet2.append(row_data)

    # 保存新表格
    i = 3
    file_directory = "output"

    for file in os.listdir(file_directory):
        if file.endswith('___table2.xlsx'):
            top_name = file[:-14]
            file_path = os.path.join(file_directory, file)
            workbook = load_workbook(file_path)

            # 获取第一个工作表
            sheet1 = workbook.active

            # 遍历 sheet1 中的每一行
            for row in sheet1.iter_rows():
                # 读取第三列的数据
                value = row[2].value

                # 将数据添加到 sheet2 的对应列
                sheet2.cell(row=row[0].row, column=i, value=value)
            sheet2.cell(row=1, column=i, value=top_name)

            i += 1

    # 保存新的工作簿
    file = os.path.join('output', 'table3.xlsx')
    workbook2.save(file)

def zong(filename):
    src_workbook = xlrd.open_workbook('NewSNP.xls')
    src_worksheet = src_workbook.sheet_by_index(0)
    hash_table2 = {}
    src_data = []
    for row in range(1, src_worksheet.nrows):  # 从第二行开始读取，跳过表头
        row_data = [cell.value for cell in src_worksheet.row(row)[:2]]
        src_data.append(row_data)
        hash_table2[row_data[1]] = []

    # 创建新的 Excel 文件并写入数据
    dst_workbook = Workbook()
    dst_worksheet = dst_workbook.active  # 使用默认的工作表

    # 添加标题行
    headers = ['snp_code', 'snp_info', 'snp顺序结果', 'snp统计结果']
    for col_index, header_text in enumerate(headers):
        col_letter = get_column_letter(col_index + 1)
        dst_cell = dst_worksheet['{}1'.format(col_letter)]
        dst_cell.value = header_text

    for row_index, row_data in enumerate(src_data):
        for col_index, cell_data in enumerate(row_data):
            dst_cell = dst_worksheet.cell(row=row_index + 2, column=col_index + 1)
            dst_cell.value = cell_data
    file_path = os.path.join('output', filename)
    with open(file_path, 'r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)  # 跳过标题行
        for row in reader:
            if len(row) >= 4:  # 确保第四列存在数据
                key = row[1]  # 第二列数据作为键
                value = row[3]  # 第四列数据作为值
                hash_table2[key].append(value)
    for row in dst_worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
        cell_value = row[0].value

        if len(hash_table2[cell_value]) != 0:
            # 键存在于 hash_table 中，将值写入对应行的第三列
            value3, value4 = sort_bases(hash_table2[cell_value])

            dst_worksheet.cell(row=row[0].row, column=3).value = value3
            dst_worksheet.cell(row=row[0].row, column=4).value = value4

        else:
            # 键不存在于 hash_table 中，对应行的第三列写入空
            dst_worksheet.cell(row=row[0].row, column=3).value = "0"
            dst_worksheet.cell(row=row[0].row, column=4).value = "A0T0G0C0"

        file = os.path.join('output', filename + '___table2.xlsx')
        dst_workbook.save(file)
    return
for filename in os.listdir('output'):
    if filename.endswith('_table1.csv'):
        zong(filename)

process_table3()
input("运行完成，按任意键结束程序")
